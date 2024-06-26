import time
from urllib.parse import urljoin

import requests
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side

# ---!!!CHANGE AIRBNB URL HERE!!!---
url = 'https://www.airbnb.com/s/Los-Angeles--California--United-States/homes?tab_id=home_tab&refinement_paths%5B%5D=%2Fhomes&flexible_trip_lengths%5B%5D=one_week&monthly_start_date=2024-05-01&monthly_length=3&monthly_end_date=2024-08-01&price_filter_input_type=0&channel=EXPLORE&query=Los%20Angeles%2C%20CA&place_id=ChIJE9on3F3HwoAR9AhGJW_fL-I&date_picker_type=calendar&checkin=2024-06-13&checkout=2024-06-15&adults=5&source=structured_search_input_header&search_type=user_map_move&price_filter_num_nights=2&ne_lat=34.099506339796726&ne_lng=-118.32489778064303&sw_lat=34.07295547099822&sw_lng=-118.35086666078894&zoom=14.524921919759125&zoom_level=14.524921919759125&search_by_map=true'

s=Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)
response = requests.get(url)

# makes sure to stop at the 18th listing on airbnb to not include the "Available for similar dates" listings
counter = 0

#EXCEL writes the headers
excelwrite = pd.DataFrame(columns=['Name', 'Listing', 'Price/Night', 'Rating', 'URL'])

while url:
    driver.get(url)
    time.sleep(2) # sleeps the page to allow it to load !!INCASE OF TIMEOUTS CHANGE THIS VALUE!!

    # wonton soup is probably my favorite
    soup = BeautifulSoup(driver.page_source, 'lxml')

    for item in soup.select('[itemprop="itemListElement"]'):
        # checks if the listing is a "Available for similar dates" listing and skips it
        similar_parent = item.find_parents(class_="f4kiyqs atm_2d_g2722k atm_gi_1lw5rbb atm_l8_5utakr atm_gi_goucad__oggzyc atm_l8_14vsfaa__oggzyc atm_gi_1lw5rbb__1v156lz atm_l8_5utakr__1v156lz dir dir-ltr")
        if similar_parent:
         continue
        # stops at the 18th listing until next page
        if counter == 18:
            break
        try:
            print('----')
            row = ['', '', '', '', '']

            #LISTING NAMES
            names = item.select('[itemprop="name"]')
            for name_element in names:
                name = name_element.get('content')
                print(name)
                row[0] = name
                counter += 1
            
            #LISTING CARDS
            listings = item.select('[class="t1jojoys atm_g3_1kw7nm4 atm_ks_15vqwwr atm_sq_1l2sidv atm_9s_cj1kg8 atm_6w_1e54zos atm_fy_1vgr820 atm_7l_18pqv07 atm_cs_qo5vgd atm_w4_1eetg7c atm_ks_zryt35__1rgatj2 dir dir-ltr"]')
            for listing_element in listings:
                listing = listing_element.text
                print(listing)
                row[1] = listing

            #PRICING/NIGHT
            prices = item.select('span._1y74zjx')
            for price_element in prices:
                price = price_element.text
                print(price + 'per night')
                row[2] = price

            #RATINGS
            ratings = item.select('[class="ru0q88m atm_cp_1ts48j8 dir dir-ltr"]')
            for rating_element in ratings:
                rating = rating_element.text
                print('★ ' + rating)
                row[3] = rating

            #LISTING URLS
            listurl = item.select('[itemprop="url"]')
            for url_element in listurl:
                urlname = url_element.get('content')
                print(urlname)
                row[4] = urlname

            #EXCEL writes to rows & columns
            excelwrite.loc[len(excelwrite)] = row

        except Exception as e:
            print(e)

    # try to push next page button if not avail then boom done
    nextpage = soup.find("a", attrs={"aria-label": "Next"})
    try:
        element = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.XPATH, '//a[@aria-label="Next"]'))
        )
    except:
        url = None
        continue

    if nextpage:
        relative_url = nextpage.get('href')
        base_url = "https://www.airbnb.com"
        url = urljoin(base_url, relative_url)
        counter = 0
        #DEBUG print ('counter set to 0')
    else:
        url = None

#EXCEL LOGIC 
excelwrite['SortPrice'] = excelwrite['Price/Night'].str.replace(r'[^\d.]', '', regex=True).astype(float) # creates new sortprice column using price/night data
excelwrite = excelwrite.sort_values(by='SortPrice', ascending=True) # sorts by sortprice which sorts price/night
excelwrite.drop(columns=['SortPrice'], inplace=True) # drops the sortprice column

darkbg = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
lightbg = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
# DISABLED - fontcolor = Font(color='EBEBEB')
borderstyle = Side(border_style="thin")
border = Border(top=borderstyle, bottom=borderstyle)
excelwb = Workbook()
excelws = excelwb.active
for r in dataframe_to_rows(excelwrite, index=False, header=True):
    excelws.append(r)

#EXCEL this fixes my hyperlinks by adding http & https if not there
for row in excelws.iter_rows(min_row=2, min_col=5, max_col=5):
    for cell in row:
        if cell.value and not cell.value.startswith(('http://', 'https://')):
            cell.value = 'https://' + cell.value
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"

#EXCEL formatting!!!
        
#EXCEL bold headers
for cell in excelws[1]:
    cell.font = Font(bold=True)

#EXCEL background & font colors
for row in excelws.iter_rows(min_row=2):
    for cell in row:
        #DISABLED - cell.font = fontcolor
        if cell.row % 2 == 0:
            cell.fill = darkbg
        else:
            cell.fill = lightbg

#EXCEL top and bottom borders
for row in excelws.iter_rows():
    for cell in row:
        cell.border = border

#EXCEL coloumn widths
excelws.column_dimensions[get_column_letter(1)].width = 500/12  # Name
excelws.column_dimensions[get_column_letter(2)].width = 300/12  # Listing
excelws.column_dimensions[get_column_letter(3)].width = 130/12  # Price
excelws.column_dimensions[get_column_letter(4)].width = 120/12  # Rating
excelws.column_dimensions[get_column_letter(5)].width = 1500/12 # URL

#EXCEL save file WOOOO
excelfilename = 'Airbnb.xlsx'
excelwb.save(excelfilename)
print('---FINISHED---')
print('saved to ' + excelfilename)

#DEBUG 
#print("Status code:", response.status_code)  
#print('--------------------------------------------------')
